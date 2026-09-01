from pathlib import Path
import csv
from openpyxl import load_workbook

src = Path('Vouchers Mini app Template.xlsx')
out = Path('extracted')
out.mkdir(exist_ok=True)

wb = load_workbook(src, data_only=True)

for ws in wb.worksheets:
    safe = ''.join(c if c.isalnum() or c in (' ', '-', '_') else '_' for c in ws.title).strip().replace(' ', '_')
    path = out / f'{safe}.csv'
    with path.open('w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            writer.writerow(['' if v is None else v for v in row])

with (out / '_sheet_index.csv').open('w', newline='', encoding='utf-8-sig') as f:
    writer = csv.writer(f)
    writer.writerow(['sheet_name','rows','columns'])
    for ws in wb.worksheets:
        writer.writerow([ws.title, ws.max_row, ws.max_column])

print('Extracted:', ', '.join(ws.title for ws in wb.worksheets))
